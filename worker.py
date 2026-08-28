#!/usr/bin/env python3
"""Worker v3.1: XLSX streaming -> COPY -> staging -> merge SQL."""

import logging
import os
import signal
import socket
import tempfile
import time
from datetime import date, datetime, time as dt_time, timezone
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote
from zoneinfo import ZoneInfo

import httpx
import psycopg
from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

DATABASE_URL = os.environ["DATABASE_URL"]
SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_SERVICE_KEY:
    raise RuntimeError("Falta SUPABASE_SERVICE_KEY o SUPABASE_SERVICE_ROLE_KEY")

BUCKET = os.environ.get("REPORTS_BUCKET", "reportes")
POLL_SECONDS = float(os.environ.get("WORKER_POLL_SECONDS", "2"))
HEARTBEAT_EVERY = int(os.environ.get("WORKER_HEARTBEAT_EVERY", "1000"))
STALE_MINUTES = int(os.environ.get("WORKER_STALE_MINUTES", "10"))
WORKER_ID = os.environ.get("WORKER_ID") or f"{socket.gethostname()}-{os.getpid()}"
SOURCE_TIMEZONE = ZoneInfo(os.environ.get("SOURCE_TIMEZONE", "America/Argentina/Buenos_Aires"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(name)s | %(message)s")
logger = logging.getLogger("reporteria-worker")
STOP = False

STAGING_COLUMNS = [
    "report_id", "fila_importacion", "prestador_nombre", "prestador_nombre_norm",
    "id_servicio_prestado", "id_orden_de_servicio", "tipo_de_servicio", "estado",
    "campana", "despachado_por", "alta_del_servicio", "pasado_al_despachador",
    "fecha_hora_asignado", "fecha_hora_envio_ok", "asigno_movil_ts", "hora_llego",
    "hora_finaliza", "hora_debe_salir_sms_demora", "hora_debe_salir_sms_mapa",
    "tiempo_demora_en_asignar", "demora_prometida", "demora_real",
    "duracion_del_servicio", "diferencia_prometida_real", "rango_demora_real",
    "con_envio_ok", "asigno_movil", "tiempo_suf_antes_ejecutar",
    "tiempo_suf_antes_finalizar", "regla_20min_antes", "llego_antes_sms_mapa",
    "finalizo_antes_sms_mapa", "asigno_movil_cuando_llego",
    "asigno_movil_cuando_finaliza", "demora_mayor_100min", "tiene_coordenadas",
    "movil_registrado", "es_programado", "fecha_programada", "hora_programada",
]

REQUIRED_HEADERS = {"IdServicioPrestado", "Prestador", "Campana", "AltaDelServicio", "DemoraPrometida", "DemoraReal"}


def on_signal(signum, frame):
    global STOP
    STOP = True
    logger.info("signal=%s; terminando despues del trabajo actual", signum)


signal.signal(signal.SIGTERM, on_signal)
signal.signal(signal.SIGINT, on_signal)


def db_connect(autocommit: bool = False):
    return psycopg.connect(DATABASE_URL, autocommit=autocommit, connect_timeout=15)


def claim_job() -> Optional[dict[str, Any]]:
    with db_connect() as conn, conn.cursor() as cur:
        cur.execute("select * from public.fn_reclamar_reporte_pendiente(%s, %s)", (WORKER_ID, STALE_MINUTES))
        row = cur.fetchone()
        if not row:
            return None
        return dict(zip([d.name for d in cur.description], row))


def update_report(report_id: str, **fields):
    if not fields:
        return
    assignments = ", ".join(f"{key} = %s" for key in fields)
    with db_connect(autocommit=True) as conn, conn.cursor() as cur:
        cur.execute(f"update public.reportes set {assignments} where id = %s", list(fields.values()) + [report_id])


def download_storage_file(storage_path: str, suffix: str) -> str:
    url = f"{SUPABASE_URL}/storage/v1/object/{BUCKET}/{quote(storage_path, safe='/')}"
    headers = {"Authorization": f"Bearer {SUPABASE_SERVICE_KEY}", "apikey": SUPABASE_SERVICE_KEY}
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    path = tmp.name
    try:
        with tmp, httpx.stream("GET", url, headers=headers, timeout=120.0) as response:
            response.raise_for_status()
            for chunk in response.iter_bytes(1024 * 1024):
                tmp.write(chunk)
        return path
    except Exception:
        if os.path.exists(path):
            os.unlink(path)
        raise


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = " ".join(str(value).split()).strip()
    return text or None


def normalize(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.upper() if text else None


def as_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def as_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def as_bool(value: Any) -> Optional[bool]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return True if value == 1 else False if value == 0 else None
    text = str(value).strip().upper()
    if text in {"SI", "SÍ", "S", "TRUE", "1", "YES"}:
        return True
    if text in {"NO", "N", "FALSE", "0"}:
        return False
    return None


def localize_source(value: datetime) -> datetime:
    return value.replace(tzinfo=SOURCE_TIMEZONE) if value.tzinfo is None else value


def as_datetime(value: Any, epoch) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return localize_source(value)
    if isinstance(value, date):
        return localize_source(datetime.combine(value, dt_time.min))
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            if isinstance(converted, datetime):
                return localize_source(converted)
            if isinstance(converted, date):
                return localize_source(datetime.combine(converted, dt_time.min))
        except Exception:
            return None
    try:
        return localize_source(date_parser.parse(str(value), dayfirst=True))
    except (ValueError, TypeError, OverflowError):
        return None


def as_date(value: Any, epoch) -> Optional[date]:
    parsed = as_datetime(value, epoch)
    return parsed.date() if parsed else None


def as_time(value: Any, epoch) -> Optional[dt_time]:
    if isinstance(value, dt_time):
        return value
    parsed = as_datetime(value, epoch)
    return parsed.time().replace(tzinfo=None) if parsed else None


def unique_headers(values) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for raw in values:
        base = clean_text(raw) or ""
        count = counts.get(base, 0)
        result.append(base if count == 0 else f"{base}.{count}")
        counts[base] = count + 1
    return result


def first_value(row: dict[str, Any], names: list[str]) -> Any:
    return next((row[name] for name in names if row.get(name) not in (None, "")), None)


def staging_row(report_id: str, line: int, row: dict[str, Any], epoch) -> tuple:
    provider = clean_text(row.get("Prestador"))
    assign_ts = first_value(row, ["AsignoMovilFechaHora", "AsignoMovilTs", "AsignoMovilTS", "AsignoMovil"])
    assign_bool = first_value(row, ["AsignoMovil2", "AsignoMovil.1", "AsignoMovil_2", "AsignoMovilBooleano"])
    return (
        report_id, line, provider, normalize(provider), as_int(row.get("IdServicioPrestado")),
        as_int(row.get("IdOrdenDeServicio")), clean_text(row.get("TipoDeServicio")), clean_text(row.get("Estado")),
        normalize(row.get("Campana")), clean_text(row.get("DespachadoPor")), as_datetime(row.get("AltaDelServicio"), epoch),
        as_datetime(row.get("PasadoAlDespachador"), epoch), as_datetime(row.get("FechaHoraAsignado"), epoch),
        as_datetime(row.get("FechaHoraEnvioOk"), epoch), as_datetime(assign_ts, epoch),
        as_datetime(row.get("HoraQueLlegoADarServicio"), epoch), as_datetime(row.get("HoraQueFinalizaServicio"), epoch),
        as_datetime(row.get("HoraDebeSalirSMSDemora"), epoch), as_datetime(row.get("HoraDebeSalirSMSMapa"), epoch),
        as_float(row.get("TiempoDemoraEnAsignar")), as_float(row.get("DemoraPrometida")), as_float(row.get("DemoraReal")),
        as_float(row.get("DuracionDelServicio")), as_float(row.get("DiferenciaEntreDemoraPrometidaYReal")),
        clean_text(row.get("RangoDemoraReal")), as_bool(row.get("ConEnvioOK")), as_bool(assign_bool),
        as_bool(row.get("TiempoSuficienteAntesEjecutar")), as_bool(row.get("TiempoSuficienteAntesFinalizar")),
        as_bool(row.get("Regla20MinAntes")), as_bool(row.get("LlegoAntesSMSMapa")), as_bool(row.get("FinalizoAntesSMSMapa")),
        as_bool(row.get("AsignoMovilCuandoLlego")), as_bool(row.get("AsignoMovilCuandoFinaliza")),
        as_bool(row.get("DemoraMayorALaPrometidaEn100Min")), as_bool(row.get("TieneCoordenadasDelCaso")),
        as_bool(row.get("MovilRegistrado")), as_bool(row.get("EsProgramado")),
        as_date(row.get("FechaProgramada"), epoch), as_time(row.get("HoraProgramada"), epoch),
    )


def process_tracking(job: dict[str, Any]):
    report_id = str(job["id"])
    temp_path: Optional[str] = None
    workbook = None
    try:
        update_report(report_id, etapa="descargando", heartbeat_at=datetime.now(timezone.utc))
        temp_path = download_storage_file(job["storage_path"], Path(job["file_name"]).suffix.lower() or ".xlsx")
        update_report(report_id, etapa="leyendo_excel", heartbeat_at=datetime.now(timezone.utc))
        workbook = load_workbook(temp_path, read_only=True, data_only=True, keep_links=False)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        if first is None:
            raise ValueError("El Excel no contiene filas")
        headers = unique_headers(first)
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

        copy_sql = "copy public.servicios_importacion_staging (" + ",".join(STAGING_COLUMNS) + ") from stdin"
        processed = 0
        min_date: Optional[date] = None
        max_date: Optional[date] = None
        update_report(report_id, etapa="cargando_staging", heartbeat_at=datetime.now(timezone.utc))
        with db_connect() as conn, conn.cursor() as cur:
            cur.execute("delete from public.servicios_importacion_staging where report_id = %s", (report_id,))
            with cur.copy(copy_sql) as copy:
                for line, values in enumerate(iterator, start=2):
                    record = staging_row(report_id, line, dict(zip(headers, values)), workbook.epoch)
                    copy.write_row(record)
                    processed += 1
                    alta = record[10]
                    if alta:
                        local_date = alta.astimezone(SOURCE_TIMEZONE).date()
                        min_date = local_date if min_date is None or local_date < min_date else min_date
                        max_date = local_date if max_date is None or local_date > max_date else max_date
                    if processed % HEARTBEAT_EVERY == 0:
                        update_report(report_id, etapa="cargando_staging", filas_procesadas=processed, heartbeat_at=datetime.now(timezone.utc))
            conn.commit()

        update_report(report_id, etapa="consolidando", filas_totales=processed, filas_procesadas=processed,
                      periodo_desde=min_date, periodo_hasta=max_date, heartbeat_at=datetime.now(timezone.utc))
        with db_connect() as conn, conn.cursor() as cur:
            cur.execute("select public.fn_consolidar_trackeo(%s)", (report_id,))
            result = cur.fetchone()[0]
            conn.commit()
        logger.info("report_id=%s finalizado result=%s", report_id, result)
    except Exception as exc:
        logger.exception("report_id=%s fallo", report_id)
        try:
            update_report(report_id, status="error", etapa="error", error_msg=str(exc)[:1000],
                          heartbeat_at=datetime.now(timezone.utc), finalizado_at=datetime.now(timezone.utc))
        except Exception:
            logger.exception("report_id=%s no se pudo registrar el error", report_id)
    finally:
        if workbook is not None:
            workbook.close()
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


def run():
    logger.info("worker_id=%s timezone=%s iniciado", WORKER_ID, SOURCE_TIMEZONE.key)
    while not STOP:
        try:
            job = claim_job()
            if not job:
                time.sleep(POLL_SECONDS)
                continue
            if (job.get("tipo_reporte") or "").lower() == "trackeo" or "trackeo" in (job.get("file_name") or "").lower():
                process_tracking(job)
            else:
                update_report(str(job["id"]), status="error", etapa="tipo_no_soportado",
                              error_msg="Worker v3.1 procesa Trackeo.", finalizado_at=datetime.now(timezone.utc),
                              heartbeat_at=datetime.now(timezone.utc))
        except Exception:
            logger.exception("error en loop del worker")
            time.sleep(POLL_SECONDS)
    logger.info("worker_id=%s detenido", WORKER_ID)


if __name__ == "__main__":
    run()
